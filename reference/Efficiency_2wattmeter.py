# -*- coding: utf-8 -*-
"""
Created on Tue Mar  1 15:19:39 2022
Updated Efficiency 2-wattmeter

@author: SMarcos
"""
import pyvisa
import time
import math

# efficiency = 1
# computed_output_pf = 2

def main():
    file_name = "DX0585F_operation.xlsx"
    rm = pyvisa.ResourceManager()
    input_pm = rm.open_resource("GPIB0::1::INSTR")
    output_pm1 = rm.open_resource("GPIB0::2::INSTR")
    output_pm2 = rm.open_resource("GPIB0::3::INSTR")
    
    # Power Meter Settings
    
    input_pm.write(":INPUT:VOLTAGE:AUTO ON")
    input_pm.write(":INPUT:CURRENT:AUTO ON")
    input_pm.write(":MEASURE:AVERAGING:TYPE LINEAR")
    input_pm.write(":MEASURE:AVERAGING:COUNT 64")
    input_pm.write(":MEASURE:AVERAGING:STATE ON")
    input_pm.write(":INPUT:MODE RMS")
    input_pm.write(":NUMERIC:FORMAT ASCII")
    input_pm.write(":NUMERIC:NORMAL:ITEM1 U,1")
    input_pm.write(":NUMERIC:NORMAL:ITEM2 I,1")
    input_pm.write(":NUMERIC:NORMAL:ITEM3 P,1")
    input_pm.write(":NUMERIC:NORMAL:ITEM4 LAMBDA,1")

    output_pm1.write(":INPUT:VOLTAGE:AUTO ON")
    output_pm1.write(":INPUT:CURRENT:AUTO ON")
    output_pm1.write(":MEASURE:AVERAGING:TYPE LINEAR")
    output_pm1.write(":MEASURE:AVERAGING:COUNT 64")
    output_pm1.write(":MEASURE:AVERAGING:STATE ON")
    output_pm1.write(":INPUT:MODE RMS")
    output_pm1.write(":NUMERIC:FORMAT ASCII")
    output_pm1.write(":NUMERIC:NORMAL:ITEM1 U,1")
    output_pm1.write(":NUMERIC:NORMAL:ITEM2 I,1")
    output_pm1.write(":NUMERIC:NORMAL:ITEM3 P,1")
    output_pm1.write(":NUMERIC:NORMAL:ITEM4 LAMBDA,1")

    output_pm2.write(":INPUT:VOLTAGE:AUTO ON")
    output_pm2.write(":INPUT:CURRENT:AUTO ON")
    output_pm2.write(":MEASURE:AVERAGING:TYPE LINEAR")
    output_pm2.write(":MEASURE:AVERAGING:COUNT 64")
    output_pm2.write(":MEASURE:AVERAGING:STATE ON")
    output_pm2.write(":INPUT:MODE RMS")
    output_pm2.write(":NUMERIC:FORMAT ASCII")
    output_pm2.write(":NUMERIC:NORMAL:ITEM1 U,1")
    output_pm2.write(":NUMERIC:NORMAL:ITEM2 I,1")
    output_pm2.write(":NUMERIC:NORMAL:ITEM3 P,1")
    output_pm2.write(":NUMERIC:NORMAL:ITEM4 LAMBDA,1")
    
    # Spreadsheet
    
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "All Data"
    ws.append(["Time(mins)", 
               "Input Voltage", "Input Current", "Input PF", "Input Power", 
               "Output Voltage 1", "Output Current 1", "Output PF 1", "Output Power 1",
               "Output Voltage 2", "Output Current 2", "Output PF 2", "Output Power 2",
               "Total Output Power", "Efficiency", 
               "Irms", "Vrms", "PF", "Power Loss", "\n"])
    ws = wb.create_sheet("Last Data Point")
    ws.append(["Input DC Voltage (V)", "Input DC Current (mA)", "Input Power (W)", 
               "I_O1P (mA)","I_O2P (mA)","I_O3P (mA)",
               "Inverter Output Power (W)","Inverter Efficiency (%)","\n"])
    
    all_data = wb['All Data'] #access sheet
    last_data = wb['Last Data Point']

    print("Time(mins)", 
          "Input Voltage", "Input Current", "Input PF", "Input Power", 
          "Output Voltage 1", "Output Current 1", "Output PF 1", "Output Power 1",
          "Output Voltage 2", "Output Current 2", "Output PF 2", "Output Power 2",
          "Total Output Power", "Efficiency", 
          "Irms", "Vrms", "PF", "Power Loss", "\n")
    
    duration = int(15) # 10 minutes
    dt = float(1.0) # 1 minute
    timer = 0 
    data_no = 1
    
    all_data.append(["Data No. " + str(data_no)])
    print("Data No.", data_no)
    time.sleep(2)
    
    #Record Data
    while (1):
    
        input_pm.write("HOLD ON")
        output_pm1.write("HOLD ON")
        output_pm2.write("HOLD ON")
            
        input_voltage = float(input_pm.query("NUM:NORM:VAL? 1").strip())
        input_current = float(input_pm.query("NUM:NORM:VAL? 2").strip())
        input_power = float(input_pm.query("NUM:NORM:VAL? 3").strip())
        input_pf = float(input_pm.query("NUM:NORM:VAL? 4").strip())
    
        output_voltage1 = float(output_pm1.query("NUM:NORM:VAL? 1").strip())
        output_current1 = float(output_pm1.query("NUM:NORM:VAL? 2").strip())
        output_power1 = float(output_pm1.query("NUM:NORM:VAL? 3").strip())
        output_pf1 = float(output_pm1.query("NUM:NORM:VAL? 4").strip())
    
        output_voltage2 = float(output_pm2.query("NUM:NORM:VAL? 1").strip())
        output_current2 = float(output_pm2.query("NUM:NORM:VAL? 2").strip())
        output_power2 = float(output_pm2.query("NUM:NORM:VAL? 3").strip())
        output_pf2 = float(output_pm2.query("NUM:NORM:VAL? 4").strip())  
        
        output_power = float(output_power1 + output_power2)    
        efficiency = float((output_power/input_power)*100.0)
        irms = float((output_current1+output_current2)/2)
        vrms = float((output_voltage1+output_voltage2)/2)
        power_loss = float(input_power -  output_power)
        computed_output_pf = float(output_power/(math.sqrt(3)*irms*vrms))
        dc_input_current = float(input_current * input_pf)
        
        all_data.append([timer, 
                         input_voltage, input_current, input_pf, input_power, 
                         output_voltage1, output_current1, output_power1, output_pf1,
                         output_voltage2, output_current2, output_power2, output_pf2,
                         output_power, efficiency, 
                         irms, vrms, computed_output_pf, power_loss, "\n"])
       
        print(timer, 
              input_voltage, input_current, input_pf, input_power, 
              output_voltage1, output_current1, output_power1, output_pf1,
              output_voltage2, output_current2, output_power2, output_pf2,
              output_power, efficiency, 
              irms, vrms, computed_output_pf, power_loss, "\n")          
                
        if timer < (duration - 1e-3):
            
            input_pm.write("HOLD OFF")
            output_pm1.write("HOLD OFF")
            output_pm2.write("HOLD OFF")
            
            time.sleep(dt * 60.0) #60.0 seconds
            timer += dt
            
        else:    
            all_data.append(["\n"])    
            last_data.append([input_voltage, dc_input_current, input_power,
                              output_current1, output_current2, "",
                              output_power, efficiency, "\n"])
            
            input_pm.write("HOLD OFF")
            output_pm1.write("HOLD OFF")
            output_pm2.write("HOLD OFF")
            
            cont = input("Press 1 for data next point, 0 to end \n")
            print("cont =", cont, "\n")
            
            if (cont == '1'):
                timer = 0
                data_no = data_no + 1
                all_data.append(["Data No." + str(data_no)])
                print("Data No. ", data_no)
                continue
                
            else:
                print("End")
                break
    
    time.sleep(2)  
    input_pm.close()
    output_pm1.close()
    output_pm2.close()    

    # Formatting and Graph
        
    from openpyxl.styles import Font   
  
    # count = range(0, 11)
    # for count in count:
    #     all_data.cell((2 + (count*13)), 1).font = Font(bold=True)
    
    count = range(0, 31)
    for count in count:
        all_data.cell((2 + (count*33)), 1).font = Font(bold=True)
        
    from openpyxl.styles import Alignment

    rows = range(1, 140)
    columns = range(1, 20)
    for row in rows:
        for col in columns:
            all_data.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
            all_data.column_dimensions[chr(64 + col)].width = 18 
            last_data.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
            last_data.column_dimensions[chr(64 + col)].width = 18
            all_data.cell(1, col).font = Font(bold=True)
            last_data.cell(1, col).font = Font(bold=True)
    
    from openpyxl.chart import(LineChart, Reference,)
    
    c1 = LineChart()
    c1.title = "Efficiency Graph"
    c1.x_axis.title = 'Output Power (W)'
    c1.y_axis.title = 'Efficiency (%)'
    
    output_power_graph = Reference (last_data, min_col = 7, min_row = 3, max_col = 7, max_row = 12)
    efficiency_graph = Reference (last_data, min_col = 8, min_row = 3, max_col = 8, max_row = 12)
    c1.add_data(efficiency_graph, titles_from_data = True)
    c1.set_categories(output_power_graph)
    last_data.add_chart(c1, "C14")
    
    c1.height = 14 
    c1.width = 20 
    wb.save(file_name) 

if __name__ == "__main__":
    main()